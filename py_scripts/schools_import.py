import sys
import psycopg2
import openpyxl
import tempfile
import urllib
import boto3
import botocore
import io

def convert_data_to_values(value):
  if isinstance(value,str):
    if value == '' or value == 'null' or value == 'NULL':
      return "null"
    return "'" + value.replace("'","''") + "'"
  elif isinstance(value,type(None)):
    return "null"
  else:
    return str(value)

def load_sql_from_file(file):
  infile = open(file, "r")
  filestr = infile.read()
  infile.close()
  return filestr

def lambda_handler(event,context):
  bucket = event['Records'][0]['s3']['bucket']['name']
  key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'],encoding = 'utf-8')
  
  #download binary data
  try:
    s3 = boto3.client('s3')
    response = s3.get_object(Bucket=bucket,Key=key)
    bindata = response['Body'].read()
  except botocore.exceptions.ClientError as e:
    if e.response['Error']['Code'] == "404":
      return None
    else:
      raise
  
  #load binary data into workbook object
  try:
    workbook_object = openpyxl.load_workbook(io.BytesIO(bindata))
    sheet_object = workbook_object.active
  except Exception as e:
    print(e)
    raise e
  
  #import to database
  value_string = ""
  for row_num in range(sheet_object.max_row):
    if row_num != 0:
      if row_num!= 1:
        value_string += ","
      value_string += "("
      for col_num in range(sheet_object.max_column):
        if col_num!=0:
          value_string += ","
        value_string += convert_data_to_values(sheet_object.cell(row=row_num+1,column=col_num+1).value)
      value_string += ")"
  value_string = value_string.encode("ascii","ignore").decode()
  
  #prepare sql
  sql = load_sql_from_file('./sql_scripts/import_schools.sql')
  sql = sql.replace("%value_string%",value_string)
  
  #connect to database
  try:
    conn = psycopg2.connect(database="schools_data")
  except:
    print("Unable to connect to database")
    raise
  cur = conn.cursor()
  
  #run query
  try:
    cur.execute(sql)
  except Exception as err:
    print("Could not Exceute sql query:",sql)
    print_psycopg2_exception(err)
    raise
  cur.close()
  conn.commit()
  conn.close()
  
  return 'Success!'

def local_tester():
  
  xlsxfile = "./data/Western Cape.xlsx"
  workbook_object = openpyxl.load_workbook(xlsxfile)
  sheet_object = workbook_object.active
  
  value_string = ""
  for row_num in range(sheet_object.max_row):
    if row_num != 0:
      if row_num!= 1:
        value_string += ","
      value_string += "("
      for col_num in range(sheet_object.max_column):
        if col_num!=0:
          value_string += ","
        value_string += convert_data_to_values(sheet_object.cell(row=row_num+1,column=col_num+1).value)
      value_string += ")"
  value_string = value_string.encode("ascii","ignore").decode()
  sql = load_sql_from_file('./sql_scripts/import_schools.sql')
  sql = sql.replace("%value_string%",value_string)
  
  outfile = open("./data/insert_test.sql","w")
  outfile.write(sql)
  outfile.close()
  
  #connect to database
  try:
    conn = psycopg2.connect(database="schools_data")
  except:
    print("Unable to connect to database")
    raise
  cur = conn.cursor()
  
  #run query
  try:
    cur.execute(sql)
  except Exception as err:
    print("Could not Exceute sql query:",sql)
    print_psycopg2_exception(err)
    raise
  cur.close()
  conn.commit()
  conn.close()
  return

"""
local_tester()
"""